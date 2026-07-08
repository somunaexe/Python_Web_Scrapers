"""
UK Skilled Worker Sponsor Filter — Tech Edition
================================================
Downloads the official Home Office Register of Licensed Sponsors,
filters for A-rated Skilled Worker sponsors in tech/software,
and outputs a clean Excel spreadsheet ready for targeted outreach.

Usage:
    pip install requests pandas openpyxl
    python filter_sponsors.py

Output:
    uk_tech_sponsors.xlsx  (in the same directory)
"""

import io
import re
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ── Configuration ─────────────────────────────────────────────────────────────

# Official Home Office CSV (updates roughly monthly — re-run to refresh)
GOV_UK_CSV_URL = (
    "https://assets.publishing.service.gov.uk/media/6a3917a2c6e94f095f3efb8c/2026-06-22_-_Worker_and_Temporary_Worker.csv"
)

# Fallback: if the URL above changes, go to:
# https://www.gov.uk/government/publications/register-of-licensed-sponsors-workers
# and grab the latest CSV link from that page.

# Tech company name keywords — catches software houses, fintechs, AI labs, etc.
TECH_KEYWORDS = [
    "software", "tech", "digital", "data", "cloud", "cyber", "ai ",
    "systems", "solutions", "computing", "network", "platform",
    "analytics", "fintech", "saas", "devops", "engineering",
    "intelligence", "labs", "laboratory", "innovation", "consulting",
    "services", "it ", "ict", "web", "app", "mobile", "game",
    "robotics", "automation", "semiconductor", "microchip", "silicon",
]

# Manually verified major tech sponsors (always include regardless of name match)
KNOWN_TECH_SPONSORS = [
    "google", "amazon", "microsoft", "meta", "apple", "oracle",
    "salesforce", "ibm", "accenture", "deloitte", "capgemini",
    "infosys", "tata", "wipro", "cognizant", "hcl", "fujitsu",
    "revolut", "wise", "monzo", "starling", "checkout", "funding circle",
    "deliveroo", "ocado", "darktrace", "graphcore", "arm ",
    "bloomberg", "refinitiv", "temenos", "finastra", "avaloq",
    "deepmind", "openai", "palantir", "snowflake", "databricks",
    "stripe", "adyen", "klarna", "transferwise", "worldpay",
    "betfair", "bet365", "sky", "bbc", "channel 4",
    "softcat", "computacenter", "bytes", "presidio",
    "jane street", "citadel", "jp morgan", "goldman sachs",
    "barclays", "hsbc", "lloyds", "natwest", "rbs",
    "astrazeneca", "glaxosmithkline", "rolls royce", "bt group",
    "vodafone", "bt ", "o2 ", "virgin media",
]

# UK tech hub cities to prioritise (will be labelled in output)
TECH_HUBS = {
    "london", "manchester", "edinburgh", "cambridge", "bristol",
    "reading", "leeds", "sheffield", "birmingham", "oxford",
    "glasgow", "brighton", "guildford", "newcastle", "nottingham",
    "cardiff", "belfast", "coventry", "milton keynes", "slough",
}

OUTPUT_FILE = "uk_tech_sponsors.xlsx"


# ── Download ───────────────────────────────────────────────────────────────────

def download_csv(url: str) -> pd.DataFrame:
    print(f"Downloading sponsor register from GOV.UK ...")
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    # The CSV uses latin-1 encoding
    df = pd.read_csv(
        io.StringIO(response.content.decode("latin-1")),
        dtype=str,
    )
    df.columns = [c.strip() for c in df.columns]
    print(f"  → {len(df):,} total sponsors loaded")
    return df


# ── Filter ─────────────────────────────────────────────────────────────────────

def is_tech(name: str) -> bool:
    name_lower = name.lower()
    for kw in KNOWN_TECH_SPONSORS:
        if kw in name_lower:
            return True
    for kw in TECH_KEYWORDS:
        if kw in name_lower:
            return True
    return False


def filter_tech_sponsors(df: pd.DataFrame) -> pd.DataFrame:
    # Column names in the GOV.UK CSV (may vary slightly — print df.columns if needed)
    # Typical columns: Organisation Name, Town/City, County, Type & Rating, Route
    col_name   = next(c for c in df.columns if "organisation" in c.lower() or "name" in c.lower())
    col_city   = next((c for c in df.columns if "town" in c.lower() or "city" in c.lower()), None)
    col_county = next((c for c in df.columns if "county" in c.lower()), None)
    col_rating = next((c for c in df.columns if "rating" in c.lower() or "type" in c.lower()), None)
    col_route  = next((c for c in df.columns if "route" in c.lower()), None)

    print(f"  Columns detected: {list(df.columns)}")

    # 1. Skilled Worker route only
    if col_route:
        df = df[df[col_route].str.contains("Skilled Worker", na=False, case=False)]
        print(f"  → {len(df):,} after filtering for Skilled Worker route")

    # 2. A-rated only (not B-rated / suspended)
    if col_rating:
        df = df[df[col_rating].str.lower().str.contains("a rating", na=False)]
        print(f"  → {len(df):,} after filtering for A-rating")

    # 3. Tech company name match
    df = df[df[col_name].apply(is_tech)]
    print(f"  → {len(df):,} after tech name filter")

    # 4. Build clean output dataframe
    clean = pd.DataFrame()
    clean["Company Name"]    = df[col_name].str.strip()
    clean["City"]            = df[col_city].str.strip()  if col_city   else "—"
    clean["County / Region"] = df[col_county].str.strip() if col_county else "—"
    clean["Rating"]          = df[col_rating].str.strip() if col_rating else "A"
    clean["Tech Hub?"]       = clean["City"].str.lower().apply(
                                   lambda x: "✓" if x in TECH_HUBS else ""
                               )
    clean["Careers Page"]    = ""   # to fill manually or via enrichment
    clean["Status"]          = ""   # e.g. "Emailed", "Replied", "Applied"
    clean["Notes"]           = ""
    clean["Date Added"]      = str(date.today())

    return clean.reset_index(drop=True)


# ── Excel Output ───────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", start_color="1A1A2E")   # dark navy
SUBHDR_FILL   = PatternFill("solid", start_color="16213E")
HUB_FILL      = PatternFill("solid", start_color="0F3460")
ALT_ROW_FILL  = PatternFill("solid", start_color="F0F4FF")
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
ACCENT_COLOR  = "E94560"   # red accent for hub tick

HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT     = Font(name="Arial", size=10)
BOLD_FONT     = Font(name="Arial", bold=True, size=10)

thin = Side(style="thin", color="D0D7E8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_excel(df: pd.DataFrame, path: str):
    df.to_excel(path, index=False, sheet_name="Tech Sponsors")

    wb = load_workbook(path)
    ws = wb["Tech Sponsors"]

    col_widths = {
        "A": 42,  # Company Name
        "B": 18,  # City
        "C": 20,  # County
        "D": 9,   # Rating
        "E": 11,  # Tech Hub?
        "F": 35,  # Careers Page
        "G": 15,  # Status
        "H": 35,  # Notes
        "I": 13,  # Date Added
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Header row
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        is_hub = ws.cell(row=row_idx, column=5).value == "✓"
        fill   = HUB_FILL if is_hub else (ALT_ROW_FILL if row_idx % 2 == 0 else WHITE_FILL)

        for cell in row:
            cell.fill      = fill
            cell.font      = Font(
                name="Arial", size=10,
                color=("FFFFFF" if is_hub else "1A1A2E"),
                bold=(cell.column == 1),
            )
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = BORDER

        # Accent the hub tick
        hub_cell = ws.cell(row=row_idx, column=5)
        hub_cell.font = Font(name="Arial", size=10, bold=True,
                             color=("E94560" if not is_hub else "00FF88"))
        hub_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 18

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Instructions sheet ─────────────────────────────────────────────────────
    wi = wb.create_sheet("How To Use")
    instructions = [
        ("UK Tech Sponsor Outreach Tracker", True, 14),
        ("", False, 11),
        ("What this file is", True, 11),
        ("A filtered extract of the UK Home Office Register of Licensed Sponsors,", False, 10),
        ("showing only A-rated Skilled Worker sponsors whose names match tech/software keywords.", False, 10),
        ("", False, 10),
        ("Column guide", True, 11),
        ("Company Name   — Official registered name (may differ from trading name)", False, 10),
        ("City           — Registered office city (may have offices elsewhere)", False, 10),
        ("County/Region  — Broader region", False, 10),
        ("Rating         — A = fully compliant, can issue new CoS immediately", False, 10),
        ("Tech Hub?      — ✓ = city is a recognised UK tech cluster (highlighted in blue)", False, 10),
        ("Careers Page   — Fill in the company's /careers URL for quick access", False, 10),
        ("Status         — Track your outreach: Researching / Emailed / Replied / Applied / Rejected", False, 10),
        ("Notes          — Any context: recruiter name, role fit, salary info, etc.", False, 10),
        ("", False, 10),
        ("How to refresh this file", True, 11),
        ("1. Go to: https://www.gov.uk/government/publications/register-of-licensed-sponsors-workers", False, 10),
        ("2. Download the latest CSV", False, 10),
        ("3. Re-run filter_sponsors.py — it will overwrite this file with fresh data", False, 10),
        ("", False, 10),
        ("Outreach strategy (direct-to-company, no public posting needed)", True, 11),
        ("Step 1 — Filter by Tech Hub? = ✓ to focus on active UK tech clusters", False, 10),
        ("Step 2 — Google '[Company Name] software engineer graduate' to find team leads on LinkedIn", False, 10),
        ("Step 3 — Email the engineering team directly (not HR): subject line matters", False, 10),
        ("Step 4 — Reference the role type you want + your stack (Python, ML, full-stack etc.)", False, 10),
        ("Step 5 — Attach a tailored 1-page CV — not your full master CV", False, 10),
        ("Step 6 — Follow up once after 7 days if no reply", False, 10),
        ("", False, 10),
        (f"Data as of: {date.today().strftime('%d %B %Y')} — re-run script to refresh", False, 9),
    ]

    wi.column_dimensions["A"].width = 90
    for r, (text, bold, size) in enumerate(instructions, start=1):
        cell = wi.cell(row=r, column=1, value=text)
        cell.font = Font(name="Arial", bold=bold, size=size,
                         color=("1A1A2E" if not bold else "0F3460"))
        cell.alignment = Alignment(wrap_text=True)
        wi.row_dimensions[r].height = 16 if not bold else 20

    wi.sheet_view.showGridLines = False

    wb.save(path)
    print(f"\n✅ Saved → {path}  ({len(df):,} companies)")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  UK Tech Sponsor Filter")
    print("=" * 55)

    df_raw  = download_csv(GOV_UK_CSV_URL)
    df_tech = filter_tech_sponsors(df_raw)
    write_excel(df_tech, OUTPUT_FILE)

    print("\nTop 20 companies by city (tech hubs first):")
    preview = df_tech.sort_values(
        ["Tech Hub?", "City", "Company Name"], ascending=[False, True, True]
    ).head(20)[["Company Name", "City", "County / Region"]]
    print(preview.to_string(index=False))
    print("\nDone. Open uk_tech_sponsors.xlsx to start your outreach.")


if __name__ == "__main__":
    main()
